"""Excel spreadsheet parser using openpyxl.

Extracts data from .xlsx files, converting sheets to structured text.
"""

from pathlib import Path
from typing import Optional, List

from loguru import logger

from src.ingestion.parsers.base_parser import BaseParser
from src.schema import DocumentType, ParsedDocument


class ExcelParser(BaseParser):
    """Parser for Microsoft Excel spreadsheets (.xlsx).
    
    Uses openpyxl to extract:
    - All sheets with data
    - Tables as formatted text
    - Metadata
    """
    
    SUPPORTED_EXTENSIONS = {".xlsx", ".xls"}
    
    def can_parse(self, file_path: str | Path) -> bool:
        """Check if file is an Excel spreadsheet."""
        return Path(file_path).suffix.lower() in self.SUPPORTED_EXTENSIONS
    
    def parse(
        self, file_path: str | Path, document_type: Optional[DocumentType] = None
    ) -> ParsedDocument:
        """Parse Excel spreadsheet.
        
        Args:
            file_path: Path to Excel file
            document_type: Optional document type override
            
        Returns:
            ParsedDocument with sheets as structured text
        """
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        logger.info(f"[ExcelParser] Parsing {file_path.name}...")
        
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError("openpyxl not installed. Install with: pip install openpyxl")
        
        wb = load_workbook(str(file_path), read_only=True, data_only=True)
        
        paragraphs = []
        tables_text = []
        headers = []
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            headers.append(f"Sheet: {sheet_name}")
            paragraphs.append(f"## {sheet_name}")
            
            sheet_rows = []
            for row in sheet.iter_rows(values_only=True):
                # Skip empty rows
                if all(cell is None for cell in row):
                    continue
                
                # Convert cells to strings
                row_text = [str(cell) if cell is not None else "" for cell in row]
                sheet_rows.append(" | ".join(row_text))
            
            if sheet_rows:
                table_text = "\n".join(sheet_rows)
                tables_text.append(table_text)
                paragraphs.extend(sheet_rows)
        
        wb.close()
        
        # Combine all text
        full_text = "\n\n".join(paragraphs)
        
        # Excel documents are typically schedules/tables
        if document_type is None:
            document_type = DocumentType.SCHEDULE
        
        logger.success(f"[ExcelParser] Parsed {file_path.name}: {len(full_text)} chars, {len(wb.sheetnames)} sheets")
        
        return ParsedDocument(
            file_path=str(file_path),
            file_name=file_path.name,
            document_type=document_type,
            content=full_text,
            paragraphs=paragraphs,
            page_count=len(wb.sheetnames),
            headers=headers,
            tables=tables_text,
            metadata={"sheets": wb.sheetnames},
        )
